from flask import Flask, render_template, request, jsonify, send_file, redirect, url_for, flash
import sqlite3
import json
import io
import csv
import re

app = Flask(__name__)
app.secret_key = 'sopk-interview-panel-2026'
DATABASE = 'interviews.db'


def get_db():
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    db = get_db()
    db.execute('''CREATE TABLE IF NOT EXISTS candidates (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        registration_date TEXT DEFAULT '',
        fio TEXT DEFAULT '',
        group_name TEXT DEFAULT '',
        phone TEXT DEFAULT '',
        contact_link TEXT DEFAULT '',
        clothing_size TEXT DEFAULT '',
        sopk_previously_worked TEXT DEFAULT '',
        sopk_period TEXT DEFAULT '',
        summer_practice TEXT DEFAULT '',
        practice_period TEXT DEFAULT '',
        desired_department TEXT DEFAULT '',
        source_info TEXT DEFAULT '',
        candidate_motivation TEXT DEFAULT '',
        interview_datetime TEXT DEFAULT '',
        eval_motivation TEXT DEFAULT '',
        eval_communication TEXT DEFAULT '',
        eval_responsibility TEXT DEFAULT '',
        eval_values TEXT DEFAULT '',
        eval_comments TEXT DEFAULT '',
        interview_done INTEGER DEFAULT 0,
        decision TEXT DEFAULT ''
    )''')
    for stmt in [
        "ALTER TABLE candidates ADD COLUMN interview_done INTEGER DEFAULT 0",
        "ALTER TABLE candidates ADD COLUMN decision TEXT DEFAULT ''",
    ]:
        try:
            db.execute(stmt)
        except Exception:
            pass
    db.commit()
    db.close()


DB_IMPORT_FIELDS = [
    'registration_date', 'fio', 'group_name', 'phone', 'contact_link',
    'clothing_size', 'sopk_previously_worked', 'sopk_period',
    'summer_practice', 'practice_period', 'desired_department',
    'source_info', 'candidate_motivation', 'interview_datetime',
]

# Ключи — подстроки нормализованных заголовков файла (без спецсимволов, в нижнем регистре).
# Поиск: key in normalized_header (substring match).
HEADER_TO_FIELD = {
    'дата регистрации':                                    'registration_date',
    'фио':                                                 'fio',
    'учебная группа':                                      'group_name',
    'номер телефона для связи':                            'phone',
    'телефон':                                             'phone',
    'ник или ссылка на max':                               'contact_link',
    'ник':                                                 'contact_link',
    'ссылка на связь':                                     'contact_link',
    'контакт для связи':                                   'contact_link',
    'размер одежды':                                       'clothing_size',
    'укажите свой размер одежды':                          'clothing_size',
    'работали ранее в сопк':                               'sopk_previously_worked',
    'работали ли ранее в сопк':                            'sopk_previously_worked',
    'период работы в сопк':                                'sopk_period',
    'временной период летом':                              'sopk_period',
    'будет ли летняя практика':                            'summer_practice',
    'в какой период будет проходить летняя практика':      'practice_period',
    'в каком направлении хотите работать':                 'desired_department',
    'откуда узнали':                                       'source_info',
    'почему вы хотите стать частью команды':               'candidate_motivation',
    'почему хотите стать частью команды':                  'candidate_motivation',
}

_DATE_HEADER_RE = re.compile(r'^\d{4}-\d{2}-\d{2}$|^\d{2}\.\d{2}\.\d{4}$')
_EMPTY_VALS = {'—', '-', '–', ''}


def _normalize(h):
    if h is None:
        return ''
    h = re.sub(r'[^\w\s]', ' ', str(h).lower(), flags=re.UNICODE)
    return re.sub(r'\s+', ' ', h).strip()


def _match_field(normalized, assigned):
    for key, field in HEADER_TO_FIELD.items():
        if field not in assigned and key in normalized:
            return field
    return None


def _build_col_map(raw_headers):
    """Returns (col_map {idx: field}, date_cols [(idx, iso_date)])."""
    col_map, date_cols, assigned = {}, [], set()
    for i, raw in enumerate(raw_headers):
        s = str(raw).strip() if raw is not None else ''
        if _DATE_HEADER_RE.match(s):
            # Normalise DD.MM.YYYY → YYYY-MM-DD
            if '.' in s:
                d, m, y = s.split('.')
                s = f'{y}-{m}-{d}'
            date_cols.append((i, s))
            continue
        field = _match_field(_normalize(raw), assigned)
        if field:
            col_map[i] = field
            assigned.add(field)
    return col_map, date_cols


def _row_to_record(row, col_map, date_cols):
    record = {f: '' for f in DB_IMPORT_FIELDS}
    for idx, field in col_map.items():
        record[field] = to_str(row[idx]) if idx < len(row) else ''
    # interview_datetime: берём первую дату-колонку с непустым значением
    for idx, _ in date_cols:
        if idx < len(row):
            val = to_str(row[idx])
            if val not in _EMPTY_VALS:
                record['interview_datetime'] = val
                break
    return record


def to_str(value):
    if value is None:
        return ''
    if hasattr(value, 'strftime'):
        return value.strftime('%Y-%m-%d %H:%M:%S')
    return str(value).strip()


@app.route('/')
def index():
    date_filter = request.args.get('date', '')
    sort_order = request.args.get('sort', 'asc')

    db = get_db()
    params = []
    where = ''
    if date_filter:
        where = "WHERE interview_datetime LIKE ?"
        params.append(f'{date_filter}%')

    order = 'ASC' if sort_order != 'desc' else 'DESC'
    candidates = db.execute(
        f'SELECT * FROM candidates {where} ORDER BY interview_datetime {order}',
        params
    ).fetchall()

    dates = db.execute(
        "SELECT DISTINCT substr(interview_datetime, 1, 10) AS date FROM candidates "
        "WHERE interview_datetime != '' AND interview_datetime IS NOT NULL ORDER BY date"
    ).fetchall()

    db.close()
    return render_template('index.html', candidates=candidates, dates=dates,
                           date_filter=date_filter, sort_order=sort_order)


@app.route('/import', methods=['POST'])
def import_file():
    if 'file' not in request.files:
        flash('Файл не выбран.', 'danger')
        return redirect(url_for('index'))

    file = request.files['file']
    if not file.filename:
        flash('Файл не выбран.', 'danger')
        return redirect(url_for('index'))

    filename = file.filename.lower()

    try:
        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            from openpyxl import load_workbook
            wb = load_workbook(file, data_only=True)
            ws = wb.active
            all_rows = [list(r) for r in ws.iter_rows(values_only=True)]
        elif filename.endswith('.csv'):
            content = file.read().decode('utf-8-sig')
            all_rows = list(csv.reader(content.splitlines()))
        else:
            flash('Поддерживаются только файлы XLSX и CSV.', 'danger')
            return redirect(url_for('index'))
    except Exception as e:
        flash(f'Ошибка при разборе файла: {e}', 'danger')
        return redirect(url_for('index'))

    if not all_rows:
        flash('Файл пуст.', 'danger')
        return redirect(url_for('index'))

    col_map, date_cols = _build_col_map(all_rows[0])
    records = [
        _row_to_record(row, col_map, date_cols)
        for row in all_rows[1:]
        if any(v is not None and str(v).strip() for v in row)
    ]

    placeholders = ', '.join(['?'] * len(DB_IMPORT_FIELDS))
    col_names = ', '.join(DB_IMPORT_FIELDS)

    db = get_db()
    db.execute('DELETE FROM candidates')
    for rec in records:
        db.execute(
            f'INSERT INTO candidates ({col_names}) VALUES ({placeholders})',
            [rec[f] for f in DB_IMPORT_FIELDS]
        )
    db.commit()
    db.close()

    flash(f'Импорт завершён: загружено {len(records)} кандидатов.', 'success')
    return redirect(url_for('index'))


@app.route('/candidate/<int:candidate_id>')
def get_candidate(candidate_id):
    db = get_db()
    c = db.execute('SELECT * FROM candidates WHERE id = ?', [candidate_id]).fetchone()
    db.close()
    if not c:
        return jsonify({'error': 'not found'}), 404
    return jsonify(dict(c))


@app.route('/candidate/<int:candidate_id>/delete', methods=['POST'])
def delete_candidate(candidate_id):
    db = get_db()
    db.execute('DELETE FROM candidates WHERE id = ?', [candidate_id])
    db.commit()
    db.close()
    return jsonify({'status': 'ok'})


@app.route('/candidate/<int:candidate_id>/save', methods=['POST'])
def save_candidate(candidate_id):
    data = request.get_json() or {}
    db = get_db()
    db.execute(
        '''UPDATE candidates SET
            registration_date = ?,
            fio = ?,
            group_name = ?,
            phone = ?,
            contact_link = ?,
            clothing_size = ?,
            sopk_previously_worked = ?,
            sopk_period = ?,
            summer_practice = ?,
            practice_period = ?,
            desired_department = ?,
            source_info = ?,
            candidate_motivation = ?,
            interview_datetime = ?,
            eval_motivation = ?,
            eval_communication = ?,
            eval_responsibility = ?,
            eval_values = ?,
            eval_comments = ?,
            interview_done = ?,
            decision = ?
           WHERE id = ?''',
        [
            data.get('registration_date', ''),
            data.get('fio', ''),
            data.get('group_name', ''),
            data.get('phone', ''),
            data.get('contact_link', ''),
            data.get('clothing_size', ''),
            data.get('sopk_previously_worked', ''),
            data.get('sopk_period', ''),
            data.get('summer_practice', ''),
            data.get('practice_period', ''),
            data.get('desired_department', ''),
            data.get('source_info', ''),
            data.get('candidate_motivation', ''),
            data.get('interview_datetime', ''),
            data.get('eval_motivation', ''),
            data.get('eval_communication', ''),
            data.get('eval_responsibility', ''),
            data.get('eval_values', ''),
            data.get('eval_comments', ''),
            1 if data.get('interview_done') else 0,
            data.get('decision', ''),
            candidate_id
        ]
    )
    db.commit()
    db.close()
    return jsonify({'status': 'ok'})


@app.route('/export/json')
def export_json():
    db = get_db()
    candidates = db.execute('SELECT * FROM candidates ORDER BY interview_datetime').fetchall()
    db.close()

    result = [
        {
            'id': c['id'],
            'registration_date': c['registration_date'] or '',
            'fio': c['fio'] or '',
            'group': c['group_name'] or '',
            'phone': c['phone'] or '',
            'contact_link': c['contact_link'] or '',
            'clothing_size': c['clothing_size'] or '',
            'sopk_previously_worked': c['sopk_previously_worked'] or '',
            'sopk_period': c['sopk_period'] or '',
            'summer_practice': c['summer_practice'] or '',
            'practice_period': c['practice_period'] or '',
            'desired_department': c['desired_department'] or '',
            'source_info': c['source_info'] or '',
            'candidate_motivation': c['candidate_motivation'] or '',
            'interview_datetime': c['interview_datetime'] or '',
            'eval_motivation': c['eval_motivation'] or '',
            'eval_communication': c['eval_communication'] or '',
            'eval_responsibility': c['eval_responsibility'] or '',
            'eval_values': c['eval_values'] or '',
            'eval_comments': c['eval_comments'] or '',
            'interview_done': bool(c['interview_done']),
            'decision': c['decision'] or '',
        }
        for c in candidates
    ]

    json_bytes = json.dumps(result, ensure_ascii=False, indent=2).encode('utf-8')
    return send_file(
        io.BytesIO(json_bytes),
        mimetype='application/json',
        as_attachment=True,
        download_name='candidates.json'
    )


@app.route('/export/xlsx')
def export_xlsx():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    db = get_db()
    candidates = db.execute('SELECT * FROM candidates ORDER BY interview_datetime').fetchall()
    db.close()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Кандидаты'

    headers = [
        'ID', 'Дата регистрации', 'ФИО', 'Учебная группа', 'Телефон',
        'Ссылка на связь', 'Размер одежды', 'Работали в СОПК', 'Период работы в СОПК',
        'Летняя практика', 'Период практики', 'Желаемое подразделение',
        'Откуда узнали', 'Мотивация кандидата', 'Дата/время собеседования',
        'Оценка: Мотивация', 'Оценка: Коммуникация', 'Оценка: Ответственность',
        'Оценка: Ценности', 'Комментарии',
        'Собеседование проведено', 'Решение'
    ]

    _decision_ru = {'yes': 'Супер, берём', 'maybe': 'Есть сомнения', 'no': 'Точно нет'}

    header_fill = PatternFill(start_color='1F3864', end_color='1F3864', fill_type='solid')

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    db_keys = [
        'id', 'registration_date', 'fio', 'group_name', 'phone',
        'contact_link', 'clothing_size', 'sopk_previously_worked',
        'sopk_period', 'summer_practice', 'practice_period',
        'desired_department', 'source_info', 'candidate_motivation',
        'interview_datetime', 'eval_motivation', 'eval_communication',
        'eval_responsibility', 'eval_values', 'eval_comments',
        'interview_done', 'decision'
    ]

    for row_idx, c in enumerate(candidates, 2):
        for col_idx, key in enumerate(db_keys, 1):
            if key == 'interview_done':
                val = 'Да' if c[key] else 'Нет'
            elif key == 'decision':
                val = _decision_ru.get(c[key] or '', '')
            else:
                val = c[key] or ''
            ws.cell(row=row_idx, column=col_idx, value=val)

    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 45)

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='candidates.xlsx'
    )


if __name__ == '__main__':
    init_db()
    app.run(debug=True, host='0.0.0.0', port=5050)
